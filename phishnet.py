import requests
import pandas as pd
import base64
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import random
import idna
import os
from datetime import datetime
import json

# Prompt the user to enter their API keys
VT_API_KEY = input("Enter your VirusTotal API key: ").strip()
RAPIDAPI_KEY = input("Enter your Domainr and Whois55 API key: ").strip()
URLSCAN_API_KEY = input("Enter your URLscan.io API key: ").strip()

# Prompt the user to enter the client name
client_name = input("Enter the client name: ").strip()

# Prompt the user to enter a comma-separated list of subdomain:TLD pairs
domains_input = input("Enter a list of comma-separated subdomain.TLD pairs (e.g., jacksonco.com, jacksonblemming.ca): ")
domains = [domain.strip() for domain in domains_input.split(',')]

# Parse the input into a dictionary
original_tlds = {}
for item in domains:
    parts = item.split('.')
    if len(parts) == 2:
        original_tlds[parts[0].strip()] = parts[1].strip()
    else:
        print(f"Invalid format for domain: {item}. Expected format: domain.TLD")

if not original_tlds:
    raise ValueError("No valid subdomain:TLD pairs provided.")

VT_RATE_LIMIT = 2  # requests per minute
VT_TIMEOUT = 30  # Timeout for API requests in seconds
MAX_RETRIES = 2
RAPIDAPI_HOST = 'domainr.p.rapidapi.com'
WHOIS55_HOST = 'whois55.p.rapidapi.com'

# Country abbreviation to full name mapping
country_mapping = {
    "DE": "Germany",
    "US": "United States",
    "CA": "Canada",
    "GB": "United Kingdom",
    "FR": "France",
    "JP": "Japan",
    "AU": "Australia",
    "CN": "China",
    "IN": "India",
    "IT": "Italy",
    "ES": "Spain",
    "VT": "Vietnam",
    "NL": "Netherlands",
    "BR": "Brazil",
    "RU": "Russia",
    "KR": "South Korea",
    "MX": "Mexico",
    "ZA": "South Africa",
    "SE": "Sweden",
    "CH": "Switzerland",
    "SG": "Singapore",
    "HK": "Hong Kong"
}

# Function to manage rate limiting
def rate_limiter():
    time.sleep(60 / VT_RATE_LIMIT)

def get_domain_hex(domain):
    encoded_domain = idna.encode(domain).decode('utf-8')
    url = f"https://dnstwister.report/api/to_hex/{encoded_domain}"
    response = requests.get(url)
    response.raise_for_status()
    return response.json()['domain_as_hexadecimal']

def get_domain_permutations(domain_hex):
    url = f"https://dnstwister.report/api/fuzz/{domain_hex}"
    response = requests.get(url)
    response.raise_for_status()
    return response.json()['fuzzy_domains']

def fetch_data_from_url(url):
    response = requests.get(url)
    response.raise_for_status()
    return response.json()

def get_country_from_ip(ip):
    url = f"https://ipinfo.io/{ip}/json"
    response = requests.get(url)
    response.raise_for_status()
    country_code = response.json().get('country', 'No DNS A Record')
    return country_mapping.get(country_code, country_code)

def get_main_domain(domain):
    parts = domain.split('.')
    if len(parts) > 2:
        return '.'.join(parts[-2:])  # Take the last two parts for the main domain
    return domain

def get_virustotal_url_status(domain):
    main_domain = get_main_domain(domain)
    for retry_count in range(MAX_RETRIES):
        try:
            url_id = base64.urlsafe_b64encode(main_domain.encode()).decode().strip("=")
            vt_url = f"https://www.virustotal.com/api/v3/urls/{url_id}"
            headers = {
                'x-apikey': VT_API_KEY
            }
            print(f"Checking URL: {vt_url}")  # Debugging statement

            response = requests.get(vt_url, headers=headers, timeout=VT_TIMEOUT)
            rate_limiter()  # Respect the rate limit

            if response.status_code == 204 or response.status_code == 429:  # Rate limit exceeded
                print("Request/day limit exceeded...")
                time.sleep(60)  # Wait 1 minute before retrying
                continue

            response.raise_for_status()
            result = response.json()
            last_analysis_stats = result.get("data", {}).get("attributes", {}).get("last_analysis_stats", {})
            if last_analysis_stats.get("malicious", 0) > 0:
                return "Possibly Malicious"
            return "Clean"
        except requests.exceptions.Timeout:
            print(f"Timeout error for domain {domain}")
            return "Timeout"
        except requests.exceptions.RequestException as e:
            print(f"Error processing domain {domain}: {e}")
            if retry_count >= MAX_RETRIES - 1:
                return "Clean"
    return "Clean"

def get_virustotal_ip_status(ip):
    for retry_count in range(MAX_RETRIES):
        try:
            vt_url = f"https://www.virustotal.com/api/v3/ip_addresses/{ip}"
            headers = {
                'x-apikey': VT_API_KEY
            }
            print(f"Checking IP: {vt_url}")  # Debugging statement

            response = requests.get(vt_url, headers=headers, timeout=VT_TIMEOUT)
            rate_limiter()  # Respect the rate limit

            if response.status_code == 204 or response.status_code == 429:  # Rate limit exceeded
                print("Request/day limit exceeded...")
                time.sleep(60)  # Wait 1 minute before retrying
                continue

            response.raise_for_status()
            result = response.json()
            last_analysis_stats = result.get("data", {}).get("attributes", {}).get("last_analysis_stats", {})
            if last_analysis_stats.get("malicious", 0) > 0:
                return "Possibly Malicious"
            return "Clean"
        except requests.exceptions.Timeout:
            print(f"Timeout error for IP {ip}")
            return "Timeout"
        except requests.exceptions.RequestException as e:
            print(f"Error processing IP {ip}: {e}")
            if retry_count >= MAX_RETRIES - 1:
                return "Clean"
    return "Clean"

def get_domain_availability_status(status):
    descriptions = {
        "active": "The website is active- check link if website is parked, offline, or online.",
        "undelegated": "The domain is registered, but the website is Offline.",
        "available": "The domain is For Sale.",
        "unavailable": "The domain is not available for registration.",
        "inactive": "The domain is registered, but the website is Offline.",
        "unknown": "The status of the domain cannot be determined.",
        "reserved": "The domain is reserved and cannot be registered by the public.",
        "deleted": "The domain has been deleted and is not currently active.",
        "on hold": "The domain registration is on hold, usually due to administrative reasons.",
        "premium": "The domain is For Sale.",
        "marketed priced active": "The domain is For Sale.",
        "active parked": "The domain is Parked.",
        "undelegated inactive": "The domain is registered, but the website is Offline."
    }
    if "marketed" in status:
        return "The domain is For Sale."
    if "reserved" in status:
        return "The domain is not available for registration"
    if "inactive" in status:
        return "The domain is registered, but the website is Offline."
    if "active" in status:
        return "The website is active- check link if website is parked, offline, or online."
    return descriptions.get(status, "Unknown status")

def check_domain_availability_status(domain):
    main_domain = '.'.join(domain.split('.')[-2:])  # Extract main domain (e.g., tool.net for enterp.tool.net)
    url = f"https://domainr.p.rapidapi.com/v2/status?domain={main_domain}"
    headers = {
        'x-rapidapi-host': RAPIDAPI_HOST,
        'x-rapidapi-key': RAPIDAPI_KEY
    }
    response = requests.get(url, headers=headers)
    response.raise_for_status()
    status = response.json()['status'][0]['status']
    return get_domain_availability_status(status)

def get_screenshot_url(domain):
    headers = {'API-Key': URLSCAN_API_KEY, 'Content-Type': 'application/json'}
    data = {"url": f"http://{domain}", "visibility": "public"}
    response = requests.post('https://urlscan.io/api/v1/scan/', headers=headers, data=json.dumps(data))
    response_data = response.json()
    uuid = response_data.get("uuid")
    if uuid:
        screenshot_url = f"https://urlscan.io/screenshots/{uuid}.png"
        return screenshot_url
    return None

def get_last_dns_change(domain):
    url = f"https://{WHOIS55_HOST}/api/v1/whois?domain={domain}"
    headers = {
        'x-rapidapi-key': RAPIDAPI_KEY,
        'x-rapidapi-host': WHOIS55_HOST
    }
    retries = 3
    timeout = 10  # seconds

    for attempt in range(retries):
        try:
            response = requests.get(url, headers=headers, timeout=timeout)
            response.raise_for_status()
            if response.status_code == 200:
                whois_data = response.json()
                updated_date_str = whois_data.get('parsed', {}).get('Updated Date', 'N/A')
                if updated_date_str != 'N/A':
                    updated_date = datetime.strptime(updated_date_str, '%Y-%m-%dT%H:%M:%SZ')
                    return updated_date.strftime('%Y-%m-%d')
                return 'N/A'
        except Timeout:
            print(f"Timeout occurred for {domain}. Retrying...")
        except HTTPError as http_err:
            print(f"HTTP error occurred for {domain}: {http_err}")
            break
        except RequestException as req_err:
            print(f"Request exception occurred for {domain}: {req_err}")
            break
        except Exception as e:
            print(f"Unexpected error occurred for {domain}: {e}")
            break
        time.sleep(2)  # Wait before retrying

    return 'N/A'

def process_domain_permutation(perm, base_subdomain, original_domain, original_tld):
    try:
        tld = perm['domain'].split('.')[-1]
        if tld in ['com', 'net', 'org', original_tld] and perm['domain'] != original_domain:
            mx_data = fetch_data_from_url(perm['has_mx_url'])

            if mx_data.get('mx', False):
                ip_data = fetch_data_from_url(perm['resolve_ip_url'])
                ip_value = ip_data.get('ip')
                if not ip_value or ip_value == False:
                    ip_value = 'No DNS A Record'
                parked_data = fetch_data_from_url(perm['parked_score_url'])
                country = get_country_from_ip(ip_value) if ip_value != 'No DNS A Record' else 'No DNS A Record'
                domain_status = get_virustotal_url_status(perm['domain'])
                ip_status = get_virustotal_ip_status(ip_value) if ip_value != 'No DNS A Record' else 'No DNS A Record'
                availability_status = check_domain_availability_status(perm['domain']) if ip_value != 'No DNS A Record' else 'The domain is registered, but the website is Offline.'
                screenshot_url = get_screenshot_url(perm['domain']) if ip_value != 'No DNS A Record' else 'No DNS A Record'

                notes = ""
                if domain_status == "Possibly Malicious" and (ip_status == "Clean" or ip_status == "Possibly Malicious" or ip_status == "No DNS A Record") and availability_status in [
                    "The website is active- check link if website is parked, offline, or online."
                ]:
                    notes = "Current Malicious Domain Behavior"
                elif domain_status == "Clean" and ip_status == "Possibly Malicious" and availability_status in [
                    "The website is active- check link if website is parked, offline, or online."
                ]:
                    notes = "Current Malicious IP Behavior"
                elif domain_status == "Possibly Malicious" and (ip_status == "Clean" or ip_status == "Possibly Malicious" or ip_status == "No DNS A Record") and availability_status in [
                    "The domain is For Sale.",
                    "The domain is Parked.",
                    "The domain is registered, but the website is Offline."
                ]:
                    notes = "Past Malicious Domain Behavior"
                elif domain_status == "Clean" and ip_status == "Possibly Malicious" and availability_status in [
                    "The domain is For Sale.",
                    "The domain is Parked.",
                    "The domain is registered, but the website is Offline."
                ]:
                    notes = "Past Malicious IP Behavior"
                else:
                    notes = "No Malicious Behavior"
                return {
                    'Domain permutation': perm['domain'],
                    'IP': ip_value,
                    'Country': country,
                    'Redirects To': parked_data.get('redirects_to', 'N/A'),
                    'Domain Reputation Status': domain_status,
                    'IP Reputation Status': ip_status,
                    'Domain Availability Status': availability_status,
                    'Website Screenshot': screenshot_url,
                    'notes': notes,
                    'Last DNS Change': get_last_dns_change(perm['domain']),  # Add last DNS change
                    'Original Domain': base_subdomain  # Temporary column to track the original domain
                }
    except Exception as e:
        print(f"Error processing permutation {perm['domain']}: {e}")
    return None

def process_subdomain(subdomain, tlds, original_tld):
    all_data = []
    permutation_count = 0
    with ThreadPoolExecutor(max_workers=2) as executor:  # Limit the number of concurrent requests
        futures = []

        # Process the original domain first
        all_tlds = tlds + [original_tld]
        for tld in all_tlds:
            domain = f"{subdomain}.{tld}"
            try:
                domain_hex = get_domain_hex(domain)
                if domain_hex:
                    permutations = get_domain_permutations(domain_hex)
                    permutation_count += len(permutations)
                    for perm in permutations:
                        futures.append(executor.submit(process_domain_permutation, perm, f"{subdomain}.{original_tld}", domain, original_tld))
            except Exception as e:
                print(f"Error processing domain {domain}: {e}")

        for future in as_completed(futures):
            result = future.result()
            if result:
                all_data.append(result)

    # Filter results to include only .com, .net, .org, and original TLD permutations
    filtered_data = [data for data in all_data if data['Domain permutation'].endswith(('.com', '.net', '.org', f".{original_tld}"))]

    return filtered_data, permutation_count

def compare_with_previous(df, client_name, today_date):
    previous_files = [f for f in os.listdir() if f.startswith(f'domain_permutations_{client_name}_')]
    previous_files.sort()
    latest_previous_file = previous_files[-1] if previous_files else None

    if latest_previous_file is None:
        print(f"No previous file found for client '{client_name}'.")
        df['New'] = 'Yes'
    else:
        print(f"Comparing with the latest previous file: '{latest_previous_file}'")
        previous_df = pd.read_excel(latest_previous_file, sheet_name='Permutations')
        if 'Domain permutation' in previous_df.columns:
            previous_domains = set(previous_df['Domain permutation'])
            df['New'] = df['Domain permutation'].apply(lambda x: 'No' if x in previous_domains else 'Yes')
        else:
            print(f"The latest previous file does not contain the 'Domain permutation' column.")
            df['New'] = 'No'

    return df

def main(subdomains, original_tlds=None):
    tlds = ['com', 'net', 'org']
    all_data = []
    permutation_summary = []
    colors = {}

    # Generate a unique color for each original domain
    for subdomain in subdomains:
        colors[subdomain] = "{:06x}".format(random.randint(0, 0xFFFFFF))

    for subdomain in subdomains:
        if '.' in subdomain:
            subdomain_name = subdomain.split('.')[0]
        else:
            subdomain_name = subdomain

        # Include original TLD in processing
        original_tld = original_tlds[subdomain]
        data, count = process_subdomain(subdomain_name, tlds, original_tld)
        all_data.extend(data)
        permutation_summary.append({
            'Domain': f"{subdomain}.{original_tlds[subdomain]}",
            'Permutations': count  # Initialize with the count of permutations
        })

    # Remove duplicates
    unique_data = list({data['Domain permutation']: data for data in all_data}.values())

    # Remove original domains from the output
    unique_data = [data for data in unique_data if f"{data['Original Domain']}" != data['Domain permutation']]

    # Create a DataFrame
    df = pd.DataFrame(unique_data)

    # Compare with the latest previous data and mark new entries
    today_date = datetime.now().strftime("%Y%m%d")
    df = compare_with_previous(df, client_name, today_date)

    # Create the summary DataFrame
    summary_df = pd.DataFrame(permutation_summary)

    # Ensure the 'Permutations' column exists before updating
    if 'Permutations' not in summary_df.columns:
        summary_df['Permutations'] = 0

    # Correctly update the number of permutations in the summary tab based on the actual data
    permutation_counts = df['Original Domain'].value_counts().reset_index(name='Permutations')
    permutation_counts.columns = ['Original Domain', 'Permutations']
    summary_df = summary_df.merge(permutation_counts, how='left', left_on='Domain', right_on='Original Domain')
    summary_df['Permutations'] = summary_df['Permutations_y'].fillna(0).astype(int)
    summary_df = summary_df.drop(columns=['Original Domain', 'Permutations_x', 'Permutations_y'])

    total_permutations = summary_df['Permutations'].sum()
    summary_df = pd.concat([summary_df, pd.DataFrame([{'Domain': 'Total', 'Permutations': total_permutations}])], ignore_index=True)

    # Get today's date in the desired format
    excel_filename = f'domain_permutations_{client_name}_{today_date}.xlsx'

    # Append the summary DataFrame to the original DataFrame
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Permutations', index=False)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

    # Load the workbook and apply coloring
    wb = load_workbook(excel_filename)
    ws = wb['Permutations']

    # Apply coloring to the permutations
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        original_domain = row[-2].value  # Get the original domain from the temporary column
        fill_color = colors.get(original_domain.split('.')[0], "FFFFFF")  # Use the subdomain part for coloring
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        for cell in row:
            cell.fill = fill

    # Remove the temporary column
    ws.delete_cols(ws.max_column-1)

    # Save the workbook
    wb.save(excel_filename)

if __name__ == "__main__":
    subdomains = [key for key in original_tlds]
    main(subdomains, original_tlds)
